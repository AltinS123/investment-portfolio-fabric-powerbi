"""
Trading212 Auto-Refresh Script
================================
Drop a new Trading212 CSV export in the same folder as this script and run:

    python trading212_refresh.py

OR specify a path:
    python trading212_refresh.py my_new_export.csv

The script will:
  1. Clean and enrich the raw CSV
  2. Rebuild all sheets in Trading212_Clean_Report.xlsx
  3. Print a summary of key stats

Requirements:
    pip install pandas openpyxl
"""

import sys
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ── CONFIG ────────────────────────────────────────────────────────────────────
DEFAULT_CSV    = 'Trading_-_export.csv'
OUTPUT_XLSX    = 'Trading212_Clean_Report.xlsx'

# ── STYLES ────────────────────────────────────────────────────────────────────
DARK_GREEN  = '1C3A2A'
MID_GREEN   = '2D6A4F'
LIGHT_GREEN = 'B7E4C7'
ACCENT      = '52B788'

HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT  = Font(name='Arial', bold=True, color=DARK_GREEN, size=14)
SUB_FONT    = Font(name='Arial', bold=True, color=MID_GREEN, size=11)
BODY_FONT   = Font(name='Arial', size=10)
BOLD_FONT   = Font(name='Arial', bold=True, size=10)
HDR_FILL    = PatternFill('solid', fgColor=MID_GREEN)
ALT_FILL    = PatternFill('solid', fgColor='F0FAF4')
YELLOW_FILL = PatternFill('solid', fgColor='FFF3CD')
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT        = Alignment(horizontal='left',   vertical='center')
RIGHT       = Alignment(horizontal='right',  vertical='center')

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols, start_col=1):
    for i, h in enumerate(cols, start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = thin_border()

def set_col_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths, start_col):
        ws.column_dimensions[get_column_letter(i)].width = w

def fmt_usd(c):
    c.number_format = '$#,##0.00'

# ── LOAD & CLEAN ──────────────────────────────────────────────────────────────
def load_and_clean(csv_path: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df['Time'] = pd.to_datetime(df['Time'])
    df['Date']     = df['Time'].dt.date
    df['Month']    = df['Time'].dt.to_period('M').astype(str)
    df['Year']     = df['Time'].dt.year
    df['Quarter']  = 'Q' + df['Time'].dt.quarter.astype(str) + ' ' + df['Time'].dt.year.astype(str)
    df['DayOfWeek']= df['Time'].dt.day_name()

    num_cols = ['No. of shares','Price / share','Exchange rate','Result','Total',
                'Withholding tax','Stamp duty reserve tax','Currency conversion fee']
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    return df

# ── AGGREGATE HELPERS ─────────────────────────────────────────────────────────
def build_ticker_summary(df):
    buys  = df[df['Action']=='Market buy']
    sells = df[df['Action']=='Market sell']
    divs  = df[df['Action']=='Dividend (Dividend)']

    buy_agg  = buys.groupby('Ticker').agg(
        Buy_Count      =('Action','count'),
        Total_Invested =('Total','sum'),
        Avg_Buy_Price  =('Price / share','mean')
    ).reset_index()
    sell_agg = sells.groupby('Ticker').agg(
        Sell_Count     =('Action','count'),
        Total_Proceeds =('Total','sum'),
        Realized_PnL   =('Result','sum'),
        Avg_Sell_Price =('Price / share','mean')
    ).reset_index()
    div_agg  = divs.groupby('Ticker').agg(
        Dividend_Income=('Total','sum')
    ).reset_index()

    ts = pd.merge(buy_agg, sell_agg, on='Ticker', how='outer').fillna(0)
    ts = pd.merge(ts, div_agg, on='Ticker', how='outer').fillna(0)
    name_map = df.dropna(subset=['Ticker']).drop_duplicates('Ticker').set_index('Ticker')['Name'].to_dict()
    ts['Name'] = ts['Ticker'].map(name_map)
    ts['Total_Return'] = ts['Realized_PnL'] + ts['Dividend_Income']
    return ts.sort_values('Realized_PnL', ascending=False)

def build_monthly(df):
    sells = df[df['Action']=='Market sell']
    buys  = df[df['Action']=='Market buy']
    divs  = df[df['Action']=='Dividend (Dividend)']
    m = sells.groupby('Month').agg(
        Sell_Proceeds=('Total','sum'),
        Realized_PnL =('Result','sum'),
        Trade_Count  =('Action','count')
    ).reset_index()
    mb = buys.groupby('Month').agg(Invested=('Total','sum')).reset_index()
    md = divs.groupby('Month').agg(Dividends=('Total','sum')).reset_index()
    m = pd.merge(m, mb, on='Month', how='outer').fillna(0)
    m = pd.merge(m, md, on='Month', how='outer').fillna(0)
    m['Total_Return'] = m['Realized_PnL'] + m['Dividends']
    return m.sort_values('Month')

# ── WRITE SHEETS ──────────────────────────────────────────────────────────────
def write_overview(wb, df, ticker_summary, monthly):
    ws = wb.active
    ws.title = '📊 Overview'
    ws.sheet_view.showGridLines = False

    deposits = df[df['Action']=='Deposit']
    sells    = df[df['Action']=='Market sell']
    buys     = df[df['Action']=='Market buy']
    dividends= df[df['Action']=='Dividend (Dividend)']

    total_dep  = deposits['Total'].sum()
    total_inv  = buys['Total'].sum()
    total_proc = sells['Total'].sum()
    total_pnl  = sells['Result'].sum()
    total_div  = dividends['Total'].sum()
    date_range = f"{df['Time'].min().strftime('%b %Y')} – {df['Time'].max().strftime('%b %Y')}"

    ws.merge_cells('A1:H1')
    ws['A1'] = '📈  TRADING 212 — Investment Portfolio Report'
    ws['A1'].font = Font(name='Arial', bold=True, size=16, color=DARK_GREEN)
    ws['A1'].fill = PatternFill('solid', fgColor='E8F5E9')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Period: {date_range}  |  Auto-generated from Trading212 Export"
    ws['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
    ws['A2'].alignment = CENTER

    kpis = [
        ('💰 Total Deposited',  f'${total_dep:,.2f}',  'FFF3CD', DARK_GREEN),
        ('📤 Total Invested',   f'${total_inv:,.2f}',  'E3F2FD', '1565C0'),
        ('📥 Total Proceeds',   f'${total_proc:,.2f}', 'E8F5E9', MID_GREEN),
        ('📈 Realized P&L',     f'${total_pnl:,.2f}',  'E0F5E9' if total_pnl>=0 else 'FFE0E0',
                                                        MID_GREEN if total_pnl>=0 else 'C00000'),
        ('🎁 Dividends',        f'${total_div:,.2f}',  'FFF8E1', '795548'),
        ('🏦 Deposits',         f'{len(deposits)} deposits', 'F8F9FA', '444444'),
        ('📊 Total Trades',     f'{len(sells)+len(buys)} trades','F8F9FA', '444444'),
        ('🏆 Winners',          f"{(sells['Result']>0).sum()}/{len(sells)} sells",'F3E5F5','6A1B9A'),
    ]
    for i, (label, val, bg, fg) in enumerate(kpis):
        col   = (i % 4) * 2 + 1
        row_l = 4 if i < 4 else 7
        row_v = 5 if i < 4 else 8
        ws.merge_cells(start_row=row_l, start_column=col, end_row=row_l, end_column=col+1)
        ws.merge_cells(start_row=row_v, start_column=col, end_row=row_v, end_column=col+1)
        cl = ws.cell(row_l, col, label)
        cv = ws.cell(row_v, col, val)
        cl.font = Font(name='Arial', size=9, color='666666')
        cv.font = Font(name='Arial', bold=True, size=13, color=fg)
        for r2 in [row_l, row_v]:
            for c2 in range(col, col+2):
                ws.cell(r2,c2).fill = PatternFill('solid', fgColor=bg)
                ws.cell(r2,c2).alignment = CENTER
                ws.cell(r2,c2).border = thin_border()

    # P&L table
    ws.merge_cells('A11:D11')
    ws['A11'] = '🏆  Realized P&L by Ticker'
    ws['A11'].font = SUB_FONT
    ws['A11'].fill = PatternFill('solid', fgColor='E8F5E9')
    header_row(ws, 12, ['Ticker','Name','Realized P&L ($)','% of Total'])

    pnl_df = ticker_summary[ticker_summary['Realized_PnL']!=0]
    for i, (_, r) in enumerate(pnl_df.iterrows(), 13):
        fill = PatternFill('solid', fgColor='F0FAF4' if i%2==0 else 'FFFFFF')
        pct  = f"{r['Realized_PnL']/total_pnl*100:.1f}%" if total_pnl else '0%'
        for j, v in enumerate([r['Ticker'], r.get('Name',''), round(r['Realized_PnL'],2), pct], 1):
            c = ws.cell(i, j, v)
            c.font  = BODY_FONT if j!=3 else Font(name='Arial',size=10,bold=True,
                       color='007700' if r['Realized_PnL']>=0 else 'CC0000')
            c.fill = fill; c.border = thin_border()
            c.alignment = RIGHT if j >= 3 else LEFT
            if j == 3: c.number_format = '$#,##0.00'

    # Monthly
    ws.merge_cells('F11:H11')
    ws['F11'] = '📅  Monthly P&L'
    ws['F11'].font = SUB_FONT
    ws['F11'].fill = PatternFill('solid', fgColor='E8F5E9')
    for j, h in enumerate(['Month','Invested ($)','P&L ($)'], 6):
        c = ws.cell(12, j, h); c.font=HEADER_FONT; c.fill=HDR_FILL
        c.alignment=CENTER; c.border=thin_border()
    for i, (_, r) in enumerate(monthly.iterrows(), 13):
        fill = PatternFill('solid', fgColor='F0FAF4' if i%2==0 else 'FFFFFF')
        for j, v in enumerate([r['Month'], round(r['Invested'],2), round(r['Realized_PnL'],2)], 6):
            c = ws.cell(i,j,v); c.font=BODY_FONT; c.fill=fill; c.border=thin_border()
            if j > 6: c.number_format = '$#,##0.00'
            if j == 8: c.font = Font(name='Arial',size=10,
                                      color='007700' if v>=0 else 'CC0000')

    set_col_widths(ws, [12,12,12,12,12,12,18,14])
    for r2 in range(4,10): ws.row_dimensions[r2].height = 22

def write_trades(wb, df):
    ws = wb.create_sheet('📋 All Trades')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    trades = df[df['Action'].isin(['Market buy','Market sell'])].copy()
    trades['Side'] = trades['Action'].map({'Market buy':'BUY','Market sell':'SELL'})

    cols = ['Date','Time','Side','Ticker','Name','No. of shares',
            'Price / share','Total','Result','Withholding tax','Month','Quarter','Year']
    labels = ['Date','DateTime','Side','Ticker','Name','Shares','Price ($)',
              'Total ($)','Realized P&L ($)','WHT ($)','Month','Quarter','Year']
    header_row(ws, 1, labels)
    for i, (_, r) in enumerate(trades[cols].iterrows(), 2):
        fill = PatternFill('solid', fgColor='FFF3F3' if r['Side']=='SELL'
                           else ('F0FAF4' if i%2==0 else 'FFFFFF'))
        for j, (col, lbl) in enumerate(zip(cols, labels), 1):
            v = r[col]
            if hasattr(v,'strftime'): v = str(v)
            elif isinstance(v, float): v = round(v, 4)
            c = ws.cell(i, j, v)
            c.font  = BODY_FONT; c.fill = fill; c.border = thin_border()
            c.alignment = RIGHT if j >= 6 else LEFT
            if '$' in lbl: c.number_format = '$#,##0.00'
            if col == 'Side':
                c.font = Font(name='Arial',size=10,bold=True,
                              color='CC0000' if v=='SELL' else '007700')
                c.alignment = CENTER
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
    set_col_widths(ws, [12,20,6,8,22,10,12,12,12,12,10,12,6])

def write_by_ticker(wb, ticker_summary):
    ws = wb.create_sheet('🏷️ By Ticker')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    cols   = ['Ticker','Name','Buy_Count','Sell_Count','Total_Invested',
              'Total_Proceeds','Realized_PnL','Dividend_Income','Total_Return',
              'Avg_Buy_Price','Avg_Sell_Price']
    labels = ['Ticker','Name','# Buys','# Sells','Invested ($)',
              'Proceeds ($)','Realized P&L ($)','Dividends ($)','Total Return ($)',
              'Avg Buy ($)','Avg Sell ($)']
    header_row(ws, 1, labels)
    for i, (_, r) in enumerate(ticker_summary.iterrows(), 2):
        fill = PatternFill('solid', fgColor='F0FAF4' if i%2==0 else 'FFFFFF')
        for j, (col, lbl) in enumerate(zip(cols, labels), 1):
            v = r.get(col,''); v = round(v,2) if isinstance(v,float) else v
            c = ws.cell(i,j,v); c.font=BODY_FONT; c.fill=fill; c.border=thin_border()
            c.alignment = RIGHT if j>4 else LEFT
            if '$' in lbl: c.number_format = '$#,##0.00'
            if col == 'Realized_PnL':
                c.font = Font(name='Arial',size=10,bold=True,
                              color='007700' if float(v or 0)>=0 else 'CC0000')
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
    set_col_widths(ws, [8,22,8,8,15,15,15,12,15,12,12])

def write_monthly(wb, monthly):
    ws = wb.create_sheet('📅 Monthly')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    cols   = ['Month','Trade_Count','Invested','Sell_Proceeds','Realized_PnL','Dividends','Total_Return']
    labels = ['Month','# Trades','Invested ($)','Proceeds ($)','Realized P&L ($)','Dividends ($)','Total Return ($)']
    header_row(ws, 1, labels)
    for i, (_, r) in enumerate(monthly[cols].iterrows(), 2):
        fill = PatternFill('solid', fgColor='F0FAF4' if i%2==0 else 'FFFFFF')
        for j, (col, lbl) in enumerate(zip(cols, labels), 1):
            v = r[col]; v = round(v,2) if isinstance(v,float) else v
            c = ws.cell(i,j,v); c.font=BODY_FONT; c.fill=fill; c.border=thin_border()
            c.alignment = RIGHT if j>1 else LEFT
            if '$' in lbl: c.number_format = '$#,##0.00'
            if col in ['Realized_PnL','Total_Return']:
                c.font = Font(name='Arial',size=10,bold=True,
                              color='007700' if float(v or 0)>=0 else 'CC0000')
    last = len(monthly)+2
    ws.cell(last,1,'TOTAL').font = BOLD_FONT
    ws.cell(last,1).fill = PatternFill('solid', fgColor='B7E4C7')
    for j in range(2, len(cols)+1):
        c = ws.cell(last, j, f'=SUM({get_column_letter(j)}2:{get_column_letter(j)}{last-1})')
        c.number_format = '$#,##0.00' if j>2 else '#,##0'
        c.font = BOLD_FONT; c.fill = PatternFill('solid', fgColor='B7E4C7')
        c.border = thin_border()
    set_col_widths(ws, [10,8,15,15,16,13,15])

def write_deposits(wb, df):
    ws = wb.create_sheet('💵 Deposits')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    deps = df[df['Action']=='Deposit'][['Date','Total','Notes']].copy()
    deps['Notes'] = deps['Notes'].fillna('').str.replace('Transaction ID: ','',regex=False)
    header_row(ws, 1, ['Date','Amount ($)','Reference / Transaction ID'])
    for i, (_, r) in enumerate(deps.iterrows(), 2):
        fill = PatternFill('solid', fgColor='FFFDE7' if i%2==0 else 'FFFFFF')
        for j, v in enumerate([str(r['Date']), round(float(r['Total']),2), str(r['Notes'])], 1):
            c = ws.cell(i,j,v); c.font=BODY_FONT; c.fill=fill; c.border=thin_border()
            if j==2: c.number_format='$#,##0.00'; c.alignment=RIGHT
            else: c.alignment=LEFT
    last = len(deps)+2
    ws.cell(last,1,'TOTAL').font = BOLD_FONT
    ws.cell(last,1).fill = PatternFill('solid', fgColor='FFF3CD')
    c2 = ws.cell(last,2,f'=SUM(B2:B{last-1})')
    c2.number_format = '$#,##0.00'; c2.font = BOLD_FONT
    c2.fill = PatternFill('solid', fgColor='FFF3CD'); c2.border = thin_border()
    set_col_widths(ws, [12,14,48])

def write_dividends(wb, df):
    ws = wb.create_sheet('🎁 Dividends')
    ws.sheet_view.showGridLines = False
    divs = df[df['Action']=='Dividend (Dividend)'].copy()
    cols   = ['Date','Ticker','Name','No. of shares','Price / share','Total','Withholding tax']
    labels = ['Date','Ticker','Name','Shares','Div/Share ($)','Amount ($)','Withholding Tax ($)']
    header_row(ws, 1, labels)
    for i, (_, r) in enumerate(divs[cols].iterrows(), 2):
        fill = PatternFill('solid', fgColor='FFF8E1' if i%2==0 else 'FFFFFF')
        for j, (col, lbl) in enumerate(zip(cols, labels), 1):
            v = r[col]
            if isinstance(v, float): v = round(v, 5 if j==5 else 2)
            c = ws.cell(i,j,v); c.font=BODY_FONT; c.fill=fill; c.border=thin_border()
            c.alignment = RIGHT if j>=4 else LEFT
            if '$' in lbl: c.number_format = '$#,##0.00000' if j==5 else '$#,##0.00'
    set_col_widths(ws, [12,8,20,12,13,12,16])

def write_powerbi(wb, df):
    ws = wb.create_sheet('⚡ PowerBI_Ready')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    trades = df[df['Action'].isin(['Market buy','Market sell'])].copy()
    trades['Side']     = trades['Action'].map({'Market buy':'BUY','Market sell':'SELL'})
    trades['Date_str'] = trades['Date'].astype(str)
    trades['Time_str'] = trades['Time'].astype(str)

    cols = ['Date_str','Time_str','Year','Month','Quarter','Side',
            'Ticker','Name','ISIN','No. of shares','Price / share',
            'Total','Result','Withholding tax']
    labels = ['Date','DateTime','Year','Month','Quarter','Side',
              'Ticker','Name','ISIN','Shares','Price_USD',
              'Total_USD','Realized_PnL','Withholding_Tax_USD']
    header_row(ws, 1, labels)
    for i, (_, r) in enumerate(trades[cols].iterrows(), 2):
        fill = PatternFill('solid', fgColor='F0FAF4' if i%2==0 else 'FFFFFF')
        for j, (col, lbl) in enumerate(zip(cols, labels), 1):
            v = r[col]
            if isinstance(v, float): v = round(v, 4)
            c = ws.cell(i,j,v); c.font=BODY_FONT; c.fill=fill; c.border=thin_border()
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
    set_col_widths(ws, [12,20,6,10,12,6,8,20,14,10,12,12,12,16])

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_CSV

    if not os.path.exists(csv_path):
        print(f"❌  File not found: {csv_path}")
        print(f"    Place your Trading212 CSV export as '{DEFAULT_CSV}' in this folder,")
        print(f"    or pass the filename as argument: python trading212_refresh.py myfile.csv")
        sys.exit(1)

    print(f"📂  Loading: {csv_path}")
    df = load_and_clean(csv_path)

    ticker_summary = build_ticker_summary(df)
    monthly        = build_monthly(df)

    wb = Workbook()
    write_overview(wb, df, ticker_summary, monthly)
    write_trades(wb, df)
    write_by_ticker(wb, ticker_summary)
    write_monthly(wb, monthly)
    write_deposits(wb, df)
    write_dividends(wb, df)
    write_powerbi(wb, df)

    out = OUTPUT_XLSX
    wb.save(out)

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    sells    = df[df['Action']=='Market sell']
    buys     = df[df['Action']=='Market buy']
    deposits = df[df['Action']=='Deposit']
    divs     = df[df['Action']=='Dividend (Dividend)']

    print(f"\n{'='*55}")
    print(f"  ✅  Trading212_Clean_Report.xlsx updated")
    print(f"{'='*55}")
    print(f"  Period       : {df['Time'].min().strftime('%b %Y')} – {df['Time'].max().strftime('%b %Y')}")
    print(f"  Rows loaded  : {len(df)}")
    print(f"  Total Deposits   : ${deposits['Total'].sum():,.2f}")
    print(f"  Total Invested   : ${buys['Total'].sum():,.2f}")
    print(f"  Total Proceeds   : ${sells['Total'].sum():,.2f}")
    print(f"  Realized P&L     : ${sells['Result'].sum():,.2f}")
    print(f"  Dividend Income  : ${divs['Total'].sum():,.2f}")
    print(f"  Total Trades     : {len(trades := df[df['Action'].isin(['Market buy','Market sell'])])}")
    print(f"  Winning Sells    : {(sells['Result']>0).sum()} / {len(sells)}")
    print(f"  Sheets created   : 7")
    print(f"{'='*55}\n")

if __name__ == '__main__':
    main()
